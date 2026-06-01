"""
Script đánh giá Latency (Tốc độ) cho các mô hình nhận diện khuôn mặt.
Các bước đo lường được cô lập hoàn toàn cho từng model để đảm bảo tính công bằng (Cache Isolated).
Kết quả được xuất trực tiếp ra file Excel (.xlsx) định dạng chuyên nghiệp.

Cách chạy:
  python benchmarks/evaluate_all_latency_excel.py --dataset dataset_clean --max-latency 200
"""
import sys, time, argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import cv2
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.face_detector import FaceDetector
from app.best_frame_selector import BestFrameSelector
from app.config import MODELS_DIR

# ── DANH SÁCH MÔ HÌNH ───────────────────────────────────────────
MODELS_TO_EVALUATE = [
    (MODELS_DIR / "face_recognition_sface_2021dec.onnx", "SFace Original", "sface"),
    (MODELS_DIR / "face_recognition_sface_2021dec_int8.onnx", "SFace INT8", "sface"),
    (MODELS_DIR / "face_recognition_sface_2021dec_int8bq.onnx", "SFace INT8 BQ", "sface"),
    (MODELS_DIR / "facelivtv2_l.onnx", "FaceLiVT v2-L", "facelivt"),
    (MODELS_DIR / "facelivtv2_l_finetuned_VN-celeb-clean.onnx", "FaceLiVT v2-L (VNCeleb)", "facelivt"),
    (MODELS_DIR / "facelivtv2_l_finetuned_dataset_clean.onnx", "FaceLiVT v2-L (DatasetClean)", "facelivt"),
    (MODELS_DIR / "facelivtv2_l_int8.onnx", "FaceLiVT v2-L INT8", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_512.onnx", "FaceLiVT v2-S", "facelivt"),
    
    # Các mô hình phụ khác
    (MODELS_DIR / "facelivtv2_m.onnx", "FaceLiVT v2-M", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_512_int8.onnx", "FaceLiVT v2-S INT8", "facelivt"),
    (MODELS_DIR / "facelivtv2_s.onnx", "FaceLiVT v2-S (112)", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_finetuned.onnx", "FaceLiVT v2-S (Finetuned)", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_vnceleb.onnx", "FaceLiVT v2-S (VNCeleb)", "facelivt"),
]

IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}

ARCFACE_DST = np.array([
    [38.2946, 51.6963], [73.5318, 51.5014],
    [56.0252, 71.7366], [41.5493, 92.3655], [70.7299, 92.2041]
], dtype=np.float32)

# ── UTILITIES ──────────────────────────────────────────────────
def imread_u(path):
    buf = np.fromfile(str(path), dtype=np.uint8)
    return cv2.imdecode(buf, cv2.IMREAD_COLOR)

def stats(vals):
    if not vals: return {"mean": 0, "med": 0, "p95": 0, "min": 0, "max": 0}
    a = np.array(vals)
    return {
        "mean": float(np.mean(a)), "med": float(np.median(a)),
        "p95": float(np.percentile(a, 95)),
        "min": float(np.min(a)), "max": float(np.max(a)),
    }

# ── WRAPPERS ───────────────────────────────────────────────────
class SFaceEmb:
    def __init__(self, model_path, name):
        self.name = name
        self.type = "sface"
        self.embed_dim = 128
        try:
            buf = np.fromfile(str(model_path), dtype=np.uint8)
            self.rec = cv2.FaceRecognizerSF.create(framework="onnx", bufferModel=buf, bufferConfig=np.array([], dtype=np.uint8))
        except TypeError:
            self.rec = cv2.FaceRecognizerSF.create(str(model_path), "")

    def align(self, frame, det):
        return self.rec.alignCrop(frame, det)

    def get_embedding(self, aligned):
        feat = self.rec.feature(aligned).flatten()
        n = np.linalg.norm(feat)
        return feat / n if n > 1e-8 else feat

class FaceLiVTEmb:
    def __init__(self, model_path, name):
        self.name = name
        self.type = "facelivt"
        import onnxruntime as ort
        self.sess = ort.InferenceSession(str(model_path), providers=['CPUExecutionProvider'])
        self.inp = self.sess.get_inputs()[0].name
        dummy = np.random.randn(1, 3, 112, 112).astype(np.float32)
        self.embed_dim = self.sess.run(None, {self.inp: dummy})[0].flatten().shape[0]

    def align(self, frame, det, size=112):
        try:
            lm = det[4:14].reshape((5, 2))
            dst = ARCFACE_DST * (float(size) / 112.0)
            M, _ = cv2.estimateAffinePartial2D(lm, dst)
            if M is None: M = cv2.getAffineTransform(lm[:3], dst[:3])
            return cv2.warpAffine(frame, M, (size, size), borderValue=0.0)
        except Exception:
            x, y, w, h = det[:4].astype(int)
            crop = frame[max(0,y):min(frame.shape[0],y+h), max(0,x):min(frame.shape[1],x+w)]
            return cv2.resize(crop, (size, size)) if crop.size > 0 else None

    def get_embedding(self, face):
        if face is None: return np.zeros(self.embed_dim, dtype=np.float32)
        rgb = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        blob = (rgb.astype(np.float32) - 127.5) / 127.5
        blob = np.transpose(blob, (2, 0, 1))[np.newaxis, ...]
        emb = self.sess.run(None, {self.inp: blob})[0].flatten()
        n = np.linalg.norm(emb)
        return emb / n if n > 1e-8 else emb

# ── BENCHMARK CORE ──────────────────────────────────────────────
def knn_match(query, matrix, labels, k=5, threshold=0.3):
    if matrix.shape[0] == 0: return None
    scores = matrix @ query
    K = min(k, len(scores))
    top_k = np.argsort(scores)[-K:][::-1]
    votes = defaultdict(int)
    for idx in top_k:
        if float(scores[idx]) >= threshold: votes[labels[idx]] += 1
    if not votes: return None
    return max(votes, key=votes.get)

def run_latency_benchmark(images, detector, selector, embedders):
    print(f"\n{'='*70}")
    print(f"⏱️ CHẠY BENCHMARK LATENCY (FULL PIPELINE CACHE-ISOLATED)")
    print(f"{'='*70}")

    print("  [+] Đang xây dựng Gallery giả lập cho KNN KNN (Warmup)...")
    galleries = {emb.name: ([], []) for emb in embedders}
    for p in images[:100]:
        img = imread_u(p)
        if img is None: continue
        box, raw = detector.detect_largest_with_raw(img)
        if raw is None: continue
        for emb in embedders:
            vec = emb.get_embedding(emb.align(img, raw))
            galleries[emb.name][0].append(vec)
            galleries[emb.name][1].append("dummy_label")
    
    for emb in embedders:
        galleries[emb.name] = (np.array(galleries[emb.name][0], dtype=np.float32), galleries[emb.name][1])

    det_times = {emb.name: [] for emb in embedders}
    bf_times = {emb.name: [] for emb in embedders}
    st_times = {emb.name: [] for emb in embedders}
    aln_times = {emb.name: [] for emb in embedders}
    emb_times = {emb.name: [] for emb in embedders}
    knn_times = {emb.name: [] for emb in embedders}
    tot_times = {emb.name: [] for emb in embedders}

    print(f"  [+] Đang đo latency độc lập trên {len(images)} ảnh cho từng model...")
    for emb in embedders:
        print(f"\n  ➤ Model: {emb.name}")
        for idx, p in enumerate(images):
            img = imread_u(p)
            if img is None: continue

            # 1. Detection
            t0 = time.perf_counter()
            box, raw = detector.detect_largest_with_raw(img)
            dt_det = (time.perf_counter() - t0) * 1000
            det_times[emb.name].append(dt_det)
            if raw is None: continue

            # 2. Best Frame
            t0 = time.perf_counter()
            selector.reset()
            selector.update(img.copy(), box, raw[4:14], raw)
            bf, br, _ = selector.get_best()
            if bf is None: bf, br = img, raw
            dt_bf = (time.perf_counter() - t0) * 1000
            bf_times[emb.name].append(dt_bf)

            # 3. Alignment
            t0 = time.perf_counter()
            aligned = emb.align(bf, br)
            dt_aln = (time.perf_counter() - t0) * 1000
            aln_times[emb.name].append(dt_aln)

            # 4. Embedding
            t0 = time.perf_counter()
            vec = emb.get_embedding(aligned)
            dt_emb = (time.perf_counter() - t0) * 1000
            emb_times[emb.name].append(dt_emb)

            # 5. KNN
            matrix, labels = galleries[emb.name]
            t0 = time.perf_counter()
            res = knn_match(vec, matrix, labels)
            dt_knn = (time.perf_counter() - t0) * 1000
            knn_times[emb.name].append(dt_knn)

            # 6. State Machine (simulate overhead)
            t0 = time.perf_counter()
            _ = res
            dt_st = (time.perf_counter() - t0) * 1000
            st_times[emb.name].append(dt_st)

            # Total
            tot_times[emb.name].append(dt_det + dt_bf + dt_aln + dt_emb + dt_knn + dt_st)

            if (idx+1) % 50 == 0 or (idx+1) == len(images):
                print(f"      ... Đã xử lý {idx+1}/{len(images)} ảnh")

    return {
        "det": {e.name: stats(det_times[e.name]) for e in embedders},
        "bf": {e.name: stats(bf_times[e.name]) for e in embedders},
        "st": {e.name: stats(st_times[e.name]) for e in embedders},
        "aln": {e.name: stats(aln_times[e.name]) for e in embedders},
        "emb": {e.name: stats(emb_times[e.name]) for e in embedders},
        "knn": {e.name: stats(knn_times[e.name]) for e in embedders},
        "tot": {e.name: stats(tot_times[e.name]) for e in embedders},
    }

def print_latency_report(res, embedders):
    short_names = [e.name.replace("FaceLiVT", "FL").replace("Original", "Orig").replace("DatasetClean", "D.Clean").replace("VNCeleb", "VN.Cel") for e in embedders]
    metrics = ['mean', 'med', 'p95', 'min', 'max']
    
    print(f"\n{'='*120}")
    print(f"  📊 BÁO CÁO TÓM TẮT LATENCY (ms) - {len(embedders)} MÔ HÌNH")
    print(f"{'='*120}")
    
    hdr = f"  {'Bước (Metric)':<18}"
    sep = f"  {'─'*18}"
    for name in short_names:
        hdr += f" │ {name[:10]:>10}"
        sep += f"─┼─{'─'*10}"
    print(hdr)
    print(sep)
        
    def print_model_step(step_name, data_key):
        for m in metrics:
            row = f"  {f'{step_name} ({m})':<18}"
            for emb in embedders:
                val = res[data_key][emb.name][m]
                row += f" │ {val:>8.2f}  "
            print(row)
        print(sep)

    print_model_step("1. Detect", "det")
    print_model_step("2. Select", "bf")
    print_model_step("3. Align", "aln")
    print_model_step("4. Embed", "emb")
    print_model_step("5. KNN", "knn")
    print_model_step("★ TOTAL", "tot")
    
    row_fps = f"  {'🚀 FPS (từ Mean)':<18}"
    for emb in embedders:
        mean_tot = res['tot'][emb.name]['mean']
        fps = 1000 / max(mean_tot, 1)
        row_fps += f" │ {fps:>8.1f}  "
    print(row_fps)
    print(f"{'='*120}")

# ── EXCEL FORMATTER ────────────────────────────────────────────
def style_excel(excel_path):
    """Định dạng file Excel trông cực kỳ chuyên nghiệp và đẹp mắt."""
    wb = openpyxl.load_workbook(excel_path)
    
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Steel Blue
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")   # Very light gray
    
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=11)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    for name in wb.sheetnames:
        ws = wb[name]
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Định dạng header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_all
        
        # 2. Định dạng body rows
        for row in range(2, ws.max_row + 1):
            row_fill = zebra_fill if row % 2 == 0 else None
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = font_body
                cell.border = border_all
                if row_fill:
                    cell.fill = row_fill
                
                col_name = ws.cell(row=1, column=col).value
                if col_name in ["Model", "Pipeline Step", "Metric"]:
                    cell.alignment = align_left
                elif col_name in ["FPS"]:
                    cell.alignment = align_right
                    cell.number_format = "0.0"
                else: # Các chỉ số thời gian latency
                    cell.alignment = align_right
                    cell.number_format = "0.00"
                    
        # 3. Tự động giãn cột phù hợp với nội dung
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(excel_path)
    print(f"📊 Đã định dạng và thiết kế bảng Excel Latency đẹp mắt thành công!")

# ── MAIN ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=str, default="dataset_clean")
    parser.add_argument("--max-latency", type=int, default=200, help="Số ảnh tối đa để chạy benchmark")
    args = parser.parse_args()

    ds = Path(args.dataset)
    if not ds.is_absolute(): ds = PROJECT_ROOT / ds
    out_dir = PROJECT_ROOT / "benchmarks"
    out_dir.mkdir(exist_ok=True)

    print("=" * 80)
    print("  🚀 CHẠY LẠI ĐÁNH GIÁ TỐC ĐỘ (LATENCY) CỦA CÁC MÔ HÌNH VÀ XUẤT EXCEL")
    print("=" * 80)

    detector = FaceDetector()
    selector = BestFrameSelector()

    embedders = []
    for path, name, mtype in MODELS_TO_EVALUATE:
        if not path.exists():
            continue
        try:
            if mtype == "sface": 
                embedders.append(SFaceEmb(path, name))
            else: 
                embedders.append(FaceLiVTEmb(path, name))
            print(f"  [+] Đã tải mô hình: {name}")
        except Exception as e:
            print(f"  [-] Lỗi tải mô hình {name}: {e}")

    if not embedders:
        print("❌ Không có mô hình nào để chạy!")
        return

    all_images = sorted(f for f in ds.rglob("*") if f.is_file() and f.suffix.lower() in IMG_EXTS)
    lat_images = all_images[:args.max_latency] if args.max_latency > 0 else all_images
    print(f"\n[+] Sử dụng {len(lat_images)} ảnh để đo Latency.")

    lat_res = run_latency_benchmark(lat_images, detector, selector, embedders)
    print_latency_report(lat_res, embedders)

    # 1. Summary DataFrame (Thời gian tổng và FPS)
    summary_data = []
    for emb in embedders:
        tot_mean = lat_res["tot"][emb.name]["mean"]
        fps = 1000 / max(tot_mean, 1)
        summary_data.append({
            "Model": emb.name,
            "Latency_Total_Mean (ms)": tot_mean,
            "Latency_Total_P95 (ms)": lat_res["tot"][emb.name]["p95"],
            "Latency_Total_Max (ms)": lat_res["tot"][emb.name]["max"],
            "Latency_Total_Min (ms)": lat_res["tot"][emb.name]["min"],
            "FPS": fps
        })
    df_summary = pd.DataFrame(summary_data)

    # 2. Detailed Steps DataFrame (Từng bước chi tiết của Pipeline)
    detailed_data = []
    steps = [
        ("1. Detection", "det"),
        ("2. Best Frame Selection", "bf"),
        ("3. Alignment", "aln"),
        ("4. Embedding Extraction", "emb"),
        ("5. KNN Matching", "knn"),
        ("★ Total Pipeline", "tot")
    ]
    for step_name, key in steps:
        for emb in embedders:
            stats_val = lat_res[key][emb.name]
            detailed_data.append({
                "Pipeline Step": step_name,
                "Model": emb.name,
                "Mean (ms)": stats_val["mean"],
                "Median (ms)": stats_val["med"],
                "P95 (ms)": stats_val["p95"],
                "Min (ms)": stats_val["min"],
                "Max (ms)": stats_val["max"]
            })
    df_detailed = pd.DataFrame(detailed_data)

    # Ghi ra file Excel
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"latency_results_{ts}.xlsx"
    excel_path = out_dir / excel_filename
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_detailed.to_excel(writer, sheet_name='Detailed Steps', index=False)

    print(f"\n✅ Đã xuất dữ liệu latency ra file: {excel_path}")
    
    style_excel(excel_path)
    
    print(f"\n🎉 Hoàn thành xuất sắc! File kết quả Latency Excel nằm tại:\n{excel_path}")

if __name__ == "__main__":
    main()
