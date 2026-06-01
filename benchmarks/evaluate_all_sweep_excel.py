"""
Script đánh giá Độ chính xác (Accuracy, FAR, FRR) cho các mô hình nhận diện khuôn mặt.
Sử dụng Threshold Sweep để tìm ra ngưỡng tối ưu nhất cho từng model.
Sử dụng face detection cache để tăng tốc độ chạy.
Kết quả được xuất ra file Excel (.xlsx) được format chuyên nghiệp.

Cách chạy:
  python benchmarks/evaluate_all_sweep_excel.py --dataset dataset_clean --test-ratio 0.2
"""
import sys, time, random, argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import cv2
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.face_detector import FaceDetector
from app.config import MODELS_DIR

# ── DANH SÁCH MÔ HÌNH ───────────────────────────────────────────
MODELS_TO_EVALUATE = [
    (MODELS_DIR / "face_recognition_sface_2021dec.onnx", "SFace Original", "sface"),
    (MODELS_DIR / "face_recognition_sface_2021dec_int8.onnx", "SFace INT8", "sface"),
    (MODELS_DIR / "face_recognition_sface_2021dec_int8bq.onnx", "SFace INT8 BQ", "sface"),
    (MODELS_DIR / "facelivtv2_l.onnx", "FaceLiVT v2-L", "facelivt"),
    (MODELS_DIR / "facelivtv2_l_finetuned_VN-celeb-clean.onnx", "FaceLiVT v2-L (VNCeleb)", "facelivt"),
    (MODELS_DIR / "facelivtv2_l_finetuned_dataset_clean.onnx", "FaceLiVT v2-L (DatasetClean)", "facelivt"),
    (MODELS_DIR / "facelivtv2_l_int8.onnx", "FaceLiVT v2-L INT8", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_512.onnx", "FaceLiVT v2-S", "facelivt"),
    
    # Các mô hình phụ khác có trong thư mục models (nếu có)
    (MODELS_DIR / "facelivtv2_m.onnx", "FaceLiVT v2-M", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_512_int8.onnx", "FaceLiVT v2-S INT8", "facelivt"),
    (MODELS_DIR / "facelivtv2_s.onnx", "FaceLiVT v2-S (112)", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_finetuned.onnx", "FaceLiVT v2-S (Finetuned)", "facelivt"),
    (MODELS_DIR / "facelivtv2_s_vnceleb.onnx", "FaceLiVT v2-S (VNCeleb)", "facelivt"),
]

IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
SEED = 42

ARCFACE_DST = np.array([
    [38.2946, 51.6963], [73.5318, 51.5014],
    [56.0252, 71.7366], [41.5493, 92.3655], [70.7299, 92.2041]
], dtype=np.float32)

# ── UTILITIES ──────────────────────────────────────────────────
def imread_u(path):
    buf = np.fromfile(str(path), dtype=np.uint8)
    return cv2.imdecode(buf, cv2.IMREAD_COLOR)

# ── WRAPPERS ───────────────────────────────────────────────────
class SFaceEmb:
    def __init__(self, model_path, name):
        self.name = name
        self.type = "sface"
        self.embed_dim = 128
        try:
            buf = np.fromfile(str(model_path), dtype=np.uint8)
            self.rec = cv2.FaceRecognizerSF.create(framework="onnx", bufferModel=buf, bufferConfig=np.array([], dtype=np.uint8))
        except TypeError:
            self.rec = cv2.FaceRecognizerSF.create(str(model_path), "")

    def align(self, frame, det):
        return self.rec.alignCrop(frame, det)

    def get_embedding(self, aligned):
        feat = self.rec.feature(aligned).flatten()
        n = np.linalg.norm(feat)
        return feat / n if n > 1e-8 else feat

class FaceLiVTEmb:
    def __init__(self, model_path, name):
        self.name = name
        self.type = "facelivt"
        import onnxruntime as ort
        self.sess = ort.InferenceSession(str(model_path), providers=['CPUExecutionProvider'])
        self.inp = self.sess.get_inputs()[0].name
        dummy = np.random.randn(1, 3, 112, 112).astype(np.float32)
        self.embed_dim = self.sess.run(None, {self.inp: dummy})[0].flatten().shape[0]

    def align(self, frame, det, size=112):
        try:
            lm = det[4:14].reshape((5, 2))
            dst = ARCFACE_DST * (float(size) / 112.0)
            M, _ = cv2.estimateAffinePartial2D(lm, dst)
            if M is None: M = cv2.getAffineTransform(lm[:3], dst[:3])
            return cv2.warpAffine(frame, M, (size, size), borderValue=0.0)
        except Exception:
            x, y, w, h = det[:4].astype(int)
            crop = frame[max(0,y):min(frame.shape[0],y+h), max(0,x):min(frame.shape[1],x+w)]
            return cv2.resize(crop, (size, size)) if crop.size > 0 else None

    def get_embedding(self, face):
        if face is None: return np.zeros(self.embed_dim, dtype=np.float32)
        rgb = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        blob = (rgb.astype(np.float32) - 127.5) / 127.5
        blob = np.transpose(blob, (2, 0, 1))[np.newaxis, ...]
        emb = self.sess.run(None, {self.inp: blob})[0].flatten()
        n = np.linalg.norm(emb)
        return emb / n if n > 1e-8 else emb

# ── ACCURACY CORE ──────────────────────────────────────────────
def split_dataset(data_dir, test_ratio=0.2):
    rng = random.Random(SEED)
    split = {}
    for pdir in sorted([d for d in data_dir.iterdir() if d.is_dir()]):
        imgs = sorted(f for f in pdir.iterdir() if f.suffix.lower() in IMG_EXTS)
        if len(imgs) < 2: continue
        rng.shuffle(imgs)
        n_test = max(1, int(len(imgs) * test_ratio))
        split[pdir.name] = {"gallery": imgs[n_test:], "probe": imgs[:n_test]}
    return split

def pre_detect_faces(split, detector):
    print("  [+] Đang chạy Face Detection một lần duy nhất và cache kết quả...")
    t0 = time.perf_counter()
    face_cache = {}
    total_imgs = 0
    detected_imgs = 0
    
    paths = []
    for label, data in split.items():
        paths.extend(data["gallery"])
        paths.extend(data["probe"])
        
    for idx, p in enumerate(paths):
        img = imread_u(p)
        total_imgs += 1
        if img is None: continue
        dets = detector.detect_all(img)
        if dets is not None:
            det = dets[int(np.argmax(dets[:, 2] * dets[:, 3]))]
            face_cache[str(p)] = det
            detected_imgs += 1
            
        if (idx + 1) % 200 == 0 or (idx + 1) == len(paths):
            print(f"      Đã phát hiện khuôn mặt: {idx+1}/{len(paths)} ảnh...")
            
    print(f"  → Đã phát hiện {detected_imgs}/{total_imgs} khuôn mặt trong {time.perf_counter()-t0:.1f}s.")
    return face_cache

def extract_all_faces(split, face_cache, embedder):
    gallery_embs, gallery_labels = [], []
    probe_embs, probe_labels = [], []

    for label, data in split.items():
        # Enroll Gallery
        for p in data["gallery"]:
            p_str = str(p)
            if p_str not in face_cache: continue
            img = imread_u(p)
            if img is None: continue
            det = face_cache[p_str]
            vec = embedder.get_embedding(embedder.align(img, det))
            if np.linalg.norm(vec) < 1e-8: continue
            gallery_embs.append(vec)
            gallery_labels.append(label)
        
        # Extract Probe
        for p in data["probe"]:
            p_str = str(p)
            if p_str not in face_cache: continue
            img = imread_u(p)
            if img is None: continue
            det = face_cache[p_str]
            vec = embedder.get_embedding(embedder.align(img, det))
            if np.linalg.norm(vec) < 1e-8: continue
            probe_embs.append(vec)
            probe_labels.append(label)

    matrix = np.array(gallery_embs, dtype=np.float32) if gallery_embs else np.zeros((0, embedder.embed_dim))
    return matrix, gallery_labels, probe_embs, probe_labels

def evaluate_at_threshold(probe_embs, probe_labels, matrix, gallery_labels, threshold):
    correct, wrong, unknown = 0, 0, 0
    for true_label, emb in zip(probe_labels, probe_embs):
        scores = matrix @ emb
        K = min(5, len(scores))
        top_k = np.argsort(scores)[-K:][::-1]
        votes = defaultdict(int)
        best_s = {}
        for idx in top_k:
            s = float(scores[idx])
            if s < threshold: continue
            c = gallery_labels[idx]
            votes[c] += 1
            if c not in best_s or s > best_s[c]: best_s[c] = s
        if not votes:
            unknown += 1
        else:
            pred = max(votes, key=lambda c: (votes[c], best_s[c]))
            if pred == true_label: correct += 1
            else: wrong += 1
            
    total = correct + wrong + unknown
    acc = correct / total * 100 if total else 0
    far = wrong / total * 100 if total else 0
    frr = unknown / total * 100 if total else 0
    return {"thr": threshold, "acc": acc, "far": far, "frr": frr, "correct": correct, "wrong": wrong, "unknown": unknown}

def run_accuracy_sweep(split, face_cache, embedders):
    print(f"\n{'='*70}")
    print(f"🎯 ĐANG CHẠY THRESHOLD SWEEP CHO CÁC MÔ HÌNH")
    print(f"{'='*70}")

    thresholds = np.arange(0.10, 0.85, 0.025).tolist()
    all_results = {}

    for i, emb in enumerate(embedders):
        print(f"\n  [{i+1}/{len(embedders)}] Đang trích xuất embeddings & quét {emb.name}...")
        t0 = time.perf_counter()
        matrix, g_labels, p_embs, p_labels = extract_all_faces(split, face_cache, emb)
        
        results = []
        for thr in thresholds:
            r = evaluate_at_threshold(p_embs, p_labels, matrix, g_labels, thr)
            results.append(r)
        
        # Chọn threshold tối ưu nhất dựa trên Accuracy cao nhất, nếu bằng nhau chọn FAR thấp nhất
        best = max(results, key=lambda r: (r["acc"], -r["far"]))
        all_results[emb.name] = {"best": best, "sweep": results}
        
        elapsed = time.perf_counter() - t0
        print(f"    -> Xong trong {elapsed:.1f}s. Gallery: {len(g_labels)}, Probes: {len(p_labels)}")
        print(f"    -> Tối ưu tại Thr={best['thr']:.3f}: Acc={best['acc']:.2f}%, FAR={best['far']:.2f}%, FRR={best['frr']:.2f}%")

    return all_results

def print_accuracy_report(acc_res, embedders):
    print(f"\n  🏆 BẢNG XẾP HẠNG ĐỘ CHÍNH XÁC (Sắp xếp theo Accuracy):")
    col_w = 28
    print(f"  {'Model':<{col_w}} │ {'Thr':>6} │ {'Acc%':>7} │ {'FAR%':>7} │ {'FRR%':>7}")
    print(f"  {'─'*col_w}─┼─{'─'*6}─┼─{'─'*7}─┼─{'─'*7}─┼─{'─'*7}")
    
    # Sắp xếp theo Accuracy giảm dần
    sorted_embs = sorted(embedders, key=lambda e: acc_res[e.name]["best"]["acc"], reverse=True)
    
    for emb in sorted_embs:
        b = acc_res[emb.name]["best"]
        print(f"  {emb.name:<{col_w}} │ {b['thr']:>6.3f} │ {b['acc']:>7.2f} │ {b['far']:>7.2f} │ {b['frr']:>7.2f}")

# ── EXCEL FORMATTER ────────────────────────────────────────────
def style_excel(excel_path):
    """Định dạng file Excel trông cực kỳ chuyên nghiệp và đẹp mắt."""
    wb = openpyxl.load_workbook(excel_path)
    
    # Định nghĩa style
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Steel Blue
    optimal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green highlight
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")   # Very light gray
    
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=11)
    font_optimal_body = Font(name=font_family, size=11, bold=True, color="375623")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    for name in wb.sheetnames:
        ws = wb[name]
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Định dạng header (dòng 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_all
        
        # 2. Định dạng body rows
        is_sweep_sheet = (name == 'Full Sweep')
        
        for row in range(2, ws.max_row + 1):
            # Kiểm tra xem dòng này có phải là optimal threshold không (cho sheet Sweep)
            is_opt = False
            if is_sweep_sheet:
                opt_val = ws.cell(row=row, column=9).value # Is_Optimal column
                is_opt = (opt_val == "YES")
                
            # Zebra striping cho dòng chẵn
            row_fill = zebra_fill if row % 2 == 0 else None
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = font_body
                cell.border = border_all
                if row_fill:
                    cell.fill = row_fill
                
                # Căn lề và định dạng số tùy theo dữ liệu cột
                col_name = ws.cell(row=1, column=col).value
                
                if col_name in ["Model", "Is_Optimal"]:
                    cell.alignment = align_left
                elif col_name in ["Threshold", "Opt_Thr"]:
                    cell.alignment = align_center
                    cell.number_format = "0.000"
                elif col_name in ["Accuracy", "FAR", "FRR"]:
                    cell.alignment = align_right
                    cell.number_format = "0.00"
                else: # Correct, Wrong, Unknown
                    cell.alignment = align_right
                    cell.number_format = "#,##0"
                
                # Tô màu highlight cho các dòng chứa Optimal Threshold
                if is_sweep_sheet and is_opt:
                    cell.fill = optimal_fill
                    if col_name == "Threshold":
                        cell.font = font_optimal_body
                        
        # 3. Tự động giãn cột phù hợp với nội dung
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            # Thêm khoảng đệm rộng hơn một chút
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(excel_path)
    print(f"📊 Đã định dạng và thiết kế bảng Excel đẹp mắt thành công!")

# ── MAIN ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=str, default="dataset_clean")
    parser.add_argument("--test-ratio", type=float, default=0.2)
    args = parser.parse_args()

    ds = Path(args.dataset)
    if not ds.is_absolute(): ds = PROJECT_ROOT / ds
    out_dir = PROJECT_ROOT / "benchmarks"
    out_dir.mkdir(exist_ok=True)

    print("=" * 80)
    print("  🚀 CHẠY LẠI ĐÁNH GIÁ THRESHOLD SWEEP CHO CÁC MÔ HÌNH VÀ XUẤT EXCEL")
    print("=" * 80)

    detector = FaceDetector()

    # Chỉ load các mô hình thực sự tồn tại
    embedders = []
    for path, name, mtype in MODELS_TO_EVALUATE:
        if not path.exists():
            continue
        try:
            if mtype == "sface": 
                embedders.append(SFaceEmb(path, name))
            else: 
                embedders.append(FaceLiVTEmb(path, name))
            print(f"  [+] Đã tải mô hình: {name}")
        except Exception as e:
            print(f"  [-] Lỗi tải mô hình {name}: {e}")

    if not embedders:
        print("❌ Không có mô hình nào để chạy!")
        return

    # Chia dữ liệu và pre-detect khuôn mặt
    split = split_dataset(ds, test_ratio=args.test_ratio)
    face_cache = pre_detect_faces(split, detector)
    
    # Quét độ chính xác qua các threshold
    acc_res = run_accuracy_sweep(split, face_cache, embedders)
    print_accuracy_report(acc_res, embedders)

    # Chuẩn bị DataFrame để xuất ra Excel
    # 1. Summary DataFrame (Best Threshold mỗi model)
    summary_data = []
    sorted_embs = sorted(embedders, key=lambda e: acc_res[e.name]["best"]["acc"], reverse=True)
    for emb in sorted_embs:
        b = acc_res[emb.name]["best"]
        summary_data.append({
            "Model": emb.name,
            "Opt_Thr": b["thr"],
            "Accuracy": b["acc"],
            "FAR": b["far"],
            "FRR": b["frr"],
            "Correct": b["correct"],
            "Wrong": b["wrong"],
            "Unknown": b["unknown"]
        })
    df_summary = pd.DataFrame(summary_data)

    # 2. Sweep DataFrame (Toàn bộ dữ liệu sweep)
    sweep_data = []
    for emb in embedders:
        best_thr = acc_res[emb.name]["best"]["thr"]
        for r in acc_res[emb.name]["sweep"]:
            is_opt = "YES" if abs(r['thr'] - best_thr) < 1e-6 else ""
            sweep_data.append({
                "Model": emb.name,
                "Threshold": r["thr"],
                "Accuracy": r["acc"],
                "FAR": r["far"],
                "FRR": r["frr"],
                "Correct": r["correct"],
                "Wrong": r["wrong"],
                "Unknown": r["unknown"],
                "Is_Optimal": is_opt
            })
    df_sweep = pd.DataFrame(sweep_data)

    # Ghi ra file Excel
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"sweep_threshold_results_{ts}.xlsx"
    excel_path = out_dir / excel_filename
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_sweep.to_excel(writer, sheet_name='Full Sweep', index=False)

    print(f"\n✅ Đã xuất dữ liệu thô ra file: {excel_path}")
    
    # Định dạng lại bảng Excel cho đẹp mắt
    style_excel(excel_path)
    
    print(f"\n🎉 Hoàn thành xuất sắc! File kết quả Excel nằm tại:\n{excel_path}")

if __name__ == "__main__":
    main()
