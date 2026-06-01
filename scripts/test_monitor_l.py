"""
Script chạy thử nghiệm quét threshold cho duy nhất mô hình FaceLiVT v2-L
kết hợp đo đạc CPU, RAM và nhiệt độ CPU trong quá trình chạy.
"""
import sys, time, argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import cv2
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.face_detector import FaceDetector
from app.config import MODELS_DIR
from scripts.system_monitor import SystemMonitor
from benchmarks.evaluate_all_sweep_excel import FaceLiVTEmb, split_dataset, pre_detect_faces, extract_all_faces, evaluate_at_threshold

def style_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Steel Blue
    optimal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green highlight
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")   # Very light gray
    
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=11)
    font_optimal_body = Font(name=font_family, size=11, bold=True, color="375623")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    for name in wb.sheetnames:
        ws = wb[name]
        ws.views.sheetView[0].showGridLines = True
        
        # Header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_all
            
        # Body
        for row in range(2, ws.max_row + 1):
            row_fill = zebra_fill if row % 2 == 0 else None
            
            is_opt = False
            if name == 'Full Sweep':
                opt_val = ws.cell(row=row, column=9).value # Is_Optimal column
                is_opt = (opt_val == "YES")
                
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = font_body
                cell.border = border_all
                if row_fill:
                    cell.fill = row_fill
                    
                col_name = ws.cell(row=1, column=col).value
                
                if col_name in ["Model", "Is_Optimal"]:
                    cell.alignment = align_left
                elif col_name in ["Threshold", "Elapsed_Seconds"]:
                    cell.alignment = align_center
                    cell.number_format = "0.000" if col_name == "Threshold" else "0.0"
                elif col_name in ["Accuracy", "FAR", "FRR", "CPU_Usage_Percent", "RAM_Usage_Percent", "CPU_Temp_C"]:
                    cell.alignment = align_right
                    cell.number_format = "0.00"
                elif col_name in ["RAM_Used_MB"]:
                    cell.alignment = align_right
                    cell.number_format = "0.0"
                else: # Correct, Wrong, Unknown
                    cell.alignment = align_right
                    cell.number_format = "#,##0"
                    
                if name == 'Full Sweep' and is_opt:
                    cell.fill = optimal_fill
                    if col_name == "Threshold":
                        cell.font = font_optimal_body
                        
        # Auto column width
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(excel_path)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=str, default="dataset_clean")
    parser.add_argument("--test-ratio", type=float, default=0.2)
    args = parser.parse_args()

    ds = Path(args.dataset)
    if not ds.is_absolute(): ds = PROJECT_ROOT / ds

    model_path = MODELS_DIR / "facelivtv2_l.onnx"
    if not model_path.exists():
        print(f"❌ Không tìm thấy mô hình FaceLiVT v2-L tại {model_path}!")
        return

    print("=" * 70)
    # 1. Khởi chạy giám sát hệ thống
    monitor = SystemMonitor(interval=0.2) # Đo mỗi 0.2s để lấy dữ liệu chi tiết hơn
    monitor.start()
    print("=" * 70)

    # Khởi tạo detector
    detector = FaceDetector()
    
    # Load mô hình L
    print(f"\n[+] Đang tải mô hình FaceLiVT v2-L...")
    t_load = time.perf_counter()
    emb = FaceLiVTEmb(model_path, "FaceLiVT v2-L")
    print(f"  -> Đã tải xong trong {time.perf_counter() - t_load:.2f}s.")

    # Chia dataset
    split = split_dataset(ds, test_ratio=args.test_ratio)
    
    # Pre-detect faces
    face_cache = pre_detect_faces(split, detector)

    # Bắt đầu trích xuất embeddings và quét threshold
    print(f"\n[+] Bắt đầu trích xuất embeddings & quét Threshold Sweep...")
    t_sweep = time.perf_counter()
    
    matrix, g_labels, p_embs, p_labels = extract_all_faces(split, face_cache, emb)
    
    thresholds = np.arange(0.10, 0.85, 0.025).tolist()
    results = []
    for thr in thresholds:
        r = evaluate_at_threshold(p_embs, p_labels, matrix, g_labels, thr)
        results.append(r)
        
    best = max(results, key=lambda r: (r["acc"], -r["far"]))
    
    elapsed = time.perf_counter() - t_sweep
    print(f"  -> Hoàn thành quét trong {elapsed:.1f}s.")
    print(f"  🏆 Kết quả tối ưu tại Thr={best['thr']:.3f}: Acc={best['acc']:.2f}%, FAR={best['far']:.2f}%, FRR={best['frr']:.2f}%")

    print("\n" + "=" * 70)
    # Dừng giám sát hệ thống và in kết quả đo đạc
    monitor.stop()

    # Chuẩn bị DataFrames
    df_sweep = pd.DataFrame([{
        "Model": "FaceLiVT v2-L",
        "Threshold": r["thr"],
        "Accuracy": r["acc"],
        "FAR": r["far"],
        "FRR": r["frr"],
        "Correct": r["correct"],
        "Wrong": r["wrong"],
        "Unknown": r["unknown"],
        "Is_Optimal": "YES" if abs(r['thr'] - best['thr']) < 1e-6 else ""
    } for r in results])
    
    df_monitor = pd.DataFrame([{
        "Elapsed_Seconds": r["elapsed_sec"],
        "CPU_Usage_Percent": r["cpu_percent"],
        "RAM_Used_MB": r["ram_used_mb"],
        "RAM_Usage_Percent": r["ram_percent"],
        "CPU_Temp_C": r["cpu_temp"]
    } for r in monitor.records])
    
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = PROJECT_ROOT / "benchmarks"
    excel_path = out_dir / f"facelivt_l_sweep_monitor_{ts}.xlsx"
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_sweep.to_excel(writer, sheet_name='Full Sweep', index=False)
        df_monitor.to_excel(writer, sheet_name='System Metrics', index=False)
        
    style_excel(excel_path)
    print(f"\n🎉 Hoàn thành! Đã lưu kết quả quét và log hệ thống vào tệp Excel:\n{excel_path}")

if __name__ == "__main__":
    main()
